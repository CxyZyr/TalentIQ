"""
候选人服务层 - 处理候选人管理的业务逻辑
"""
import asyncio
import logging
import os
import threading
import time
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from sqlalchemy import or_, func, select
from concurrent.futures import ThreadPoolExecutor, as_completed

from db.models import (
    Candidate, CandidateStageHistory, JobDescription, User, SalaryNegotiation,
    UserRole, CandidateStage, CandidateStageResult, DepartmentModel, get_china_time
)
from db.salary_negotiation_queries import (
    get_latest_salary_negotiation_record,
    get_latest_salary_negotiation_subquery,
)
from utils.resume_parser import ResumeParser
from utils.candidate_analysis_service import (
    BasicInfoExtractionService,
    SummaryExtractionService,
    HardRequirementsAssessmentService,
    AIScoreService
)

logger = logging.getLogger(__name__)


class CandidateService:
    """候选人服务类"""

    def __init__(self, db: Session):
        self.db = db
        self.resume_parser = ResumeParser()

        # 延迟初始化AI服务
        self._basic_info_service = None
        self._summary_service = None
        self._hard_req_service = None
        self._ai_score_service = None

    def _generate_candidate_number(self, offset: int = 0) -> str:
        """
        生成候选人序号

        格式：YYYYMM-XXXX
        示例：202601-0001, 202601-0002
        每月从0001开始累加，下月重新从0001开始

        Args:
            offset: 偏移量，用于批量添加时避免编号冲突
        """
        from datetime import datetime

        # 获取当前年月
        now = datetime.now()
        year_month = now.strftime("%Y%m")

        # 查询当月最大序号
        max_number = self.db.query(Candidate.candidate_number).filter(
            Candidate.candidate_number.like(f"{year_month}-%")
        ).order_by(Candidate.candidate_number.desc()).first()

        if max_number and max_number[0]:
            # 提取序号部分并加1
            last_seq = int(max_number[0].split("-")[1])
            new_seq = last_seq + 1 + offset
        else:
            # 当月第一个候选人
            new_seq = 1 + offset

        # 格式化为4位数字
        return f"{year_month}-{new_seq:04d}"

    @property
    def basic_info_service(self):
        """延迟初始化基础信息提取服务"""
        if self._basic_info_service is None:
            self._basic_info_service = BasicInfoExtractionService()
        return self._basic_info_service

    @property
    def summary_service(self):
        """延迟初始化概况提取服务"""
        if self._summary_service is None:
            self._summary_service = SummaryExtractionService()
        return self._summary_service

    @property
    def hard_req_service(self):
        """延迟初始化硬性条件评估服务"""
        if self._hard_req_service is None:
            self._hard_req_service = HardRequirementsAssessmentService()
        return self._hard_req_service

    @property
    def ai_score_service(self):
        """延迟初始化AI评分服务"""
        if self._ai_score_service is None:
            self._ai_score_service = AIScoreService()
        return self._ai_score_service

    def add_candidates(
        self,
        jd_id: int,
        resume_files: List[str],
        screening_owner_id: int,
        user_id: int
    ) -> Dict[str, Any]:
        """
        添加候选人（批量）

        Args:
            jd_id: 所属JD ID
            resume_files: 简历文件路径列表
            screening_owner_id: 简历筛选负责人ID
            user_id: 创建人ID（HR）

        Returns:
            添加结果
        """
        # 1. 权限检查：只有HR可以添加候选人
        user = self.db.query(User).filter(User.id == user_id).first()
        if not user or user.role != UserRole.HR.value:
            raise ValueError("只有HR可以添加候选人")

        # 2. 检查JD是否存在
        jd = self.db.query(JobDescription).filter(JobDescription.id == jd_id).first()
        if not jd:
            raise ValueError("JD不存在")

        # 3. 检查简历筛选负责人是否存在
        screening_owner = self.db.query(User).filter(User.id == screening_owner_id).first()
        if not screening_owner:
            raise ValueError("简历筛选负责人不存在")

        # 4. 解析所有简历文件
        parse_results = self.resume_parser.parse_multiple_resumes(resume_files)

        # 5. 为每个成功解析的简历创建候选人记录
        candidates = []
        for idx, (file_path, parse_result) in enumerate(parse_results.items()):
            if parse_result['success']:
                # 生成候选人序号，传入偏移量避免批量添加时编号冲突
                candidate_number = self._generate_candidate_number(offset=len(candidates))

                candidate = Candidate(
                    candidate_number=candidate_number,
                    jd_id=jd_id,
                    created_by=user_id,
                    resume_file_path=file_path,
                    resume_text=parse_result['text'],
                    current_stage=CandidateStage.RESUME_SCREENING.value,
                    current_stage_result=CandidateStageResult.PENDING.value,
                    current_stage_owner=screening_owner_id,
                    is_parsed=False
                )
                self.db.add(candidate)
                candidates.append(candidate)

        # 6. 提交到数据库，获取候选人ID
        self.db.commit()
        for candidate in candidates:
            self.db.refresh(candidate)

        # 7. 待办将在简历解析完成后创建（见 _process_single_candidate）

        # 8. 异步处理候选人分析（后台任务）
        # 使用线程池并发处理多个候选人
        candidate_ids = [c.id for c in candidates]
        self._process_candidates_async(candidate_ids, jd)

        return {
            "total": len(resume_files),
            "success": len(candidates),
            "failed": len(resume_files) - len(candidates),
            "candidate_ids": candidate_ids,
            "parse_errors": {
                file_path: result['error']
                for file_path, result in parse_results.items()
                if not result['success']
            }
        }

    def _process_candidates_async(self, candidate_ids: List[int], jd: JobDescription):
        """
        异步处理候选人分析（后台任务）

        在后台线程中执行，不阻塞主请求

        Args:
            candidate_ids: 候选人ID列表
            jd: JD对象
        """
        # 提取JD信息，避免在线程中使用ORM对象
        jd_info = {
            'id': jd.id,
            'job_title': jd.job_title,
            'job_responsibilities': jd.job_responsibilities,
            'hard_requirements': jd.hard_requirements,
            'extracted_hard_requirements': jd.extracted_hard_requirements,
            'resume_rule_set_id': jd.resume_rule_set_id
        }

        def background_task():
            """后台执行的任务"""
            logger.info("[候选人解析] 后台批处理开始 candidate_ids=%s jd_id=%s", candidate_ids, jd_info["id"])
            # 使用线程池并发处理多个候选人
            with ThreadPoolExecutor(max_workers=3) as executor:
                futures = []
                for candidate_id in candidate_ids:
                    future = executor.submit(
                        self._process_single_candidate,
                        candidate_id,
                        jd_info
                    )
                    futures.append(future)

                # 等待所有任务完成
                for future in as_completed(futures):
                    try:
                        future.result()
                    except Exception as e:
                        logger.exception("[候选人解析] 候选人处理失败 error=%s", str(e))

            logger.info("[候选人解析] 后台批处理结束 candidate_ids=%s jd_id=%s", candidate_ids, jd_info["id"])

        # 在后台线程中执行，不阻塞主请求
        thread = threading.Thread(target=background_task, daemon=True)
        thread.start()
        logger.info("[候选人解析] 后台线程已启动 candidate_ids=%s thread_name=%s", candidate_ids, thread.name)

    def _run_analysis_step(self, candidate_id: int, step_name: str, func, *args):
        """执行单个解析步骤并记录日志"""
        start_time = time.time()
        logger.info("[候选人解析] 步骤开始 candidate_id=%s step=%s", candidate_id, step_name)
        try:
            result = func(*args)
            logger.info(
                "[候选人解析] 步骤成功 candidate_id=%s step=%s elapsed=%.2fs",
                candidate_id,
                step_name,
                time.time() - start_time,
            )
            return result
        except Exception as e:
            logger.exception(
                "[候选人解析] 步骤失败 candidate_id=%s step=%s elapsed=%.2fs error=%s",
                candidate_id,
                step_name,
                time.time() - start_time,
                str(e),
            )
            raise

    def _process_single_candidate(self, candidate_id: int, jd_info: Dict[str, Any]):
        """
        处理单个候选人的分析任务

        Args:
            candidate_id: 候选人ID
            jd_info: JD信息字典
        """
        # 创建新的数据库会话（线程安全）
        from db.database import SessionLocal
        db = SessionLocal()
        total_start = time.time()

        try:
            # 获取候选人信息
            candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
            if not candidate:
                logger.warning("[候选人解析] 候选人不存在 candidate_id=%s", candidate_id)
                return

            logger.info(
                "[候选人解析] 开始 candidate_id=%s candidate_number=%s jd_id=%s resume_text_len=%s",
                candidate.id,
                candidate.candidate_number,
                candidate.jd_id,
                len(candidate.resume_text or ""),
            )

            resume_text = candidate.resume_text

            # 准备JD信息
            jd_data = {
                'job_title': jd_info['job_title'],
                'job_responsibilities': jd_info['job_responsibilities'],
                'hard_requirements': jd_info['hard_requirements']
            }

            # 获取评分规则（根据规则集ID查询所有规则）
            from db.models import ResumeEvaluationRule
            evaluation_rules = db.query(ResumeEvaluationRule).filter(
                ResumeEvaluationRule.rule_set_id == jd_info['resume_rule_set_id']
            ).all() if jd_info.get('resume_rule_set_id') else []

            # 使用线程池并行执行4个AI分析任务
            with ThreadPoolExecutor(max_workers=4) as executor:
                # 提交4个任务
                future_basic_info = executor.submit(
                    self._run_analysis_step,
                    candidate_id,
                    "basic_info",
                    self.basic_info_service.extract_basic_info,
                    resume_text
                )
                future_summary = executor.submit(
                    self._run_analysis_step,
                    candidate_id,
                    "summary",
                    self.summary_service.extract_summary,
                    resume_text,
                    jd_data
                )
                future_hard_req = executor.submit(
                    self._run_analysis_step,
                    candidate_id,
                    "hard_requirements",
                    self.hard_req_service.assess_hard_requirements,
                    resume_text,
                    jd_info.get('extracted_hard_requirements') or {}
                )
                future_ai_score = executor.submit(
                    self._run_analysis_step,
                    candidate_id,
                    "ai_score",
                    self.ai_score_service.score_resume,
                    resume_text,
                    evaluation_rules
                )

                # 等待所有任务完成并获取结果
                basic_info_result = future_basic_info.result()
                summary_result = future_summary.result()
                hard_req_result = future_hard_req.result()
                ai_score_result = future_ai_score.result()

            # 更新候选人信息 - 基础信息
            if 'parse_error' not in basic_info_result:
                candidate.name = basic_info_result.get('name')
                candidate.gender = basic_info_result.get('gender')
                candidate.age = basic_info_result.get('age')
                candidate.work_status = basic_info_result.get('work_status')
                candidate.work_years = basic_info_result.get('work_years')
                candidate.expected_salary = basic_info_result.get('expected_salary')
                candidate.highest_education = basic_info_result.get('highest_education')
                candidate.school = basic_info_result.get('school')
                candidate.basic_info_json = basic_info_result

                # 评估学校等级（985/211/双一流）
                if candidate.school:
                    self._evaluate_school(candidate, db)

            db.commit()

            # 更新候选人信息 - 基本概况
            candidate.summary = summary_result
            db.commit()

            # 更新候选人信息 - 硬性条件评估
            if 'parse_error' not in hard_req_result:
                candidate.hard_requirements_assessment = hard_req_result
                candidate.hard_requirements_passed = hard_req_result.get('overall_passed', False)
            db.commit()

            # 更新候选人信息 - AI评分
            if 'parse_error' not in ai_score_result:
                candidate.ai_score_detail = ai_score_result
                candidate.ai_score_main = ai_score_result.get('main_score', 0)
                candidate.ai_score_bonus = ai_score_result.get('bonus_score', 0)
                candidate.ai_score_total = ai_score_result.get('total_score', 0)
            db.commit()

            # 标记为解析完成
            candidate.is_parsed = True
            db.commit()
            logger.info(
                "[候选人解析] 成功 candidate_id=%s candidate_number=%s total_elapsed=%.2fs",
                candidate.id,
                candidate.candidate_number,
                time.time() - total_start,
            )

            # 简历解析完成后创建待办并发送邮件通知
            try:
                from services.todo_service import TodoService
                from db.models import CandidateStage
                TodoService.create_todo(
                    db=db,
                    candidate_id=candidate_id,
                    stage=CandidateStage.RESUME_SCREENING.value,
                    owner_id=candidate.current_stage_owner
                )
                TodoService._send_todo_email(db, candidate_id, CandidateStage.RESUME_SCREENING.value, candidate.current_stage_owner)
            except Exception as e:
                logger.exception("[简历筛选待办/邮件] 发送失败 candidate_id=%s error=%s", candidate_id, str(e))

        except Exception as e:
            # 记录错误
            candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
            if candidate:
                candidate.parse_error = str(e)
                candidate.is_parsed = True  # 即使失败也标记为已处理
                db.commit()
                logger.exception(
                    "[候选人解析] 失败 candidate_id=%s candidate_number=%s total_elapsed=%.2fs error=%s",
                    candidate.id,
                    candidate.candidate_number,
                    time.time() - total_start,
                    str(e),
                )
            else:
                logger.exception(
                    "[候选人解析] 失败且候选人不存在 candidate_id=%s total_elapsed=%.2fs error=%s",
                    candidate_id,
                    time.time() - total_start,
                    str(e),
                )
        finally:
            db.close()

    def _evaluate_school(self, candidate: Candidate, db: Session):
        """
        评估候选人学校等级（985/211/双一流）

        Args:
            candidate: 候选人对象
            db: 数据库会话
        """
        from db.models import School

        school_name = candidate.school.strip()

        # 在学校表中查找匹配的学校
        school = db.query(School).filter(School.name == school_name).first()

        if school:
            # 找到匹配的学校，更新候选人信息
            candidate.school_id = school.id
            candidate.is_985 = school.is_985
            candidate.is_211 = school.is_211
            candidate.is_double_first_class = school.is_double_first_class
        else:
            # 未找到匹配的学校，尝试模糊匹配
            # 去除常见的后缀（如"大学"、"学院"等）
            school_name_clean = school_name.replace('大学', '').replace('学院', '').strip()

            # 模糊匹配
            school = db.query(School).filter(
                School.name.like(f'%{school_name_clean}%')
            ).first()

            if school:
                candidate.school_id = school.id
                candidate.is_985 = school.is_985
                candidate.is_211 = school.is_211
                candidate.is_double_first_class = school.is_double_first_class
            else:
                # 完全未找到，标记为非985/211/双一流
                candidate.school_id = None
                candidate.is_985 = False
                candidate.is_211 = False
                candidate.is_double_first_class = False

    def _get_user_by_id(self, user_id: int) -> User:
        """获取用户对象"""
        user = self.db.query(User).filter(User.id == user_id).first()
        if not user:
            raise ValueError("用户不存在")
        return user

    def _can_user_access_candidate(self, candidate: Candidate, user: User) -> bool:
        """判断用户是否有权限查看候选人"""
        if user.role in [UserRole.HR.value, UserRole.CEO.value]:
            return True

        if user.role != UserRole.INTERVIEWER.value:
            return False

        from db.models import CandidateStageHistory

        if candidate.current_stage_owner == user.id:
            return True

        history = self.db.query(CandidateStageHistory).filter(
            CandidateStageHistory.candidate_id == candidate.id,
            or_(
                CandidateStageHistory.stage_owner == user.id,
                CandidateStageHistory.next_stage_owner == user.id
            )
        ).first()
        return history is not None

    def get_candidate_by_id(self, candidate_id: int, user_id: int) -> Optional[Candidate]:
        """
        根据ID获取候选人详情

        Args:
            candidate_id: 候选人ID
            user_id: 用户ID

        Returns:
            候选人对象
        """
        user = self._get_user_by_id(user_id)

        candidate = self.db.query(Candidate).filter(Candidate.id == candidate_id).first()
        if not candidate:
            return None

        if not self._can_user_access_candidate(candidate, user):
            raise PermissionError("无权限查看此候选人")

        return candidate

    def delete_candidate(self, candidate_id: int, user_id: int) -> bool:
        """
        删除候选人

        Args:
            candidate_id: 候选人ID
            user_id: 用户ID

        Returns:
            是否删除成功
        """
        from db.models import (
            CandidateStageHistory, InterviewQuestion, InterviewRecording,
            CandidateTodo, InterviewEvaluation, SalaryNegotiation, TalentPool
        )

        # 权限检查：只有HR可以删除候选人
        user = self.db.query(User).filter(User.id == user_id).first()
        if not user or user.role != UserRole.HR.value:
            raise ValueError("只有HR可以删除候选人")

        candidate = self.db.query(Candidate).filter(Candidate.id == candidate_id).first()
        if not candidate:
            raise ValueError("候选人不存在")

        # 人才储备库仍引用该候选人时，禁止删除，避免出现悬空关联或历史串号
        in_talent_pool = self.db.query(TalentPool).filter(
            TalentPool.candidate_id == candidate_id
        ).first()
        if in_talent_pool:
            raise ValueError("候选人已在人才储备库中，请先从人才储备库移除后再删除")

        # 先删除所有关联记录
        self.db.query(CandidateStageHistory).filter(CandidateStageHistory.candidate_id == candidate_id).delete()
        self.db.query(InterviewQuestion).filter(InterviewQuestion.candidate_id == candidate_id).delete()
        self.db.query(InterviewRecording).filter(InterviewRecording.candidate_id == candidate_id).delete()
        self.db.query(CandidateTodo).filter(CandidateTodo.candidate_id == candidate_id).delete()
        self.db.query(InterviewEvaluation).filter(InterviewEvaluation.candidate_id == candidate_id).delete()
        self.db.query(SalaryNegotiation).filter(SalaryNegotiation.candidate_id == candidate_id).delete()

        # 删除候选人
        self.db.delete(candidate)
        self.db.commit()

        return True

    def update_candidate_basic_info(
        self,
        candidate_id: int,
        user_id: int,
        update_data: Dict[str, Any]
    ) -> Candidate:
        """
        更新候选人基本信息

        Args:
            candidate_id: 候选人ID
            user_id: 用户ID
            update_data: 更新的数据

        Returns:
            更新后的候选人对象
        """
        # 权限检查：只有HR可以编辑候选人基本信息
        user = self.db.query(User).filter(User.id == user_id).first()
        if not user or user.role != UserRole.HR.value:
            raise ValueError("只有HR可以编辑候选人基本信息")

        candidate = self.db.query(Candidate).filter(Candidate.id == candidate_id).first()
        if not candidate:
            raise ValueError("候选人不存在")

        # 更新允许编辑的字段
        allowed_fields = [
            'name', 'gender', 'age', 'work_status', 'work_years',
            'expected_salary', 'highest_education', 'school', 'privacy_info'
        ]

        for field in allowed_fields:
            if field in update_data:
                setattr(candidate, field, update_data[field])

        # 如果更新了学校，重新评估学校等级
        if 'school' in update_data and update_data['school']:
            self._evaluate_school(candidate, self.db)

        self.db.commit()
        self.db.refresh(candidate)

        return candidate

    def query_candidates(
        self,
        user_id: int,
        keyword: Optional[str] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        jd_id: Optional[int] = None,
        department: Optional[str] = None,
        stage: Optional[str] = None,
        screening_result: Optional[str] = None,
        first_interview_result: Optional[str] = None,
        second_interview_result: Optional[str] = None,
        third_interview_result: Optional[str] = None,
        offer_status: Optional[str] = None,
        # 多选筛选参数（逗号分隔）
        jd_ids: Optional[str] = None,
        departments: Optional[str] = None,
        stages: Optional[str] = None,
        screening_results: Optional[str] = None,
        first_interview_results: Optional[str] = None,
        second_interview_results: Optional[str] = None,
        third_interview_results: Optional[str] = None,
        offer_statuses: Optional[str] = None,
        page: int = 1,
        page_size: int = 20
    ) -> Dict[str, Any]:
        """
        查询候选人列表

        Args:
            user_id: 用户ID
            keyword: 关键词（搜索候选人姓名/职位）
            start_date: 开始时间
            end_date: 结束时间
            jd_id: JD ID
            department: 所属部门
            stage: 流转进展
            screening_result: 简历筛选结果
            first_interview_result: 一面结果
            second_interview_result: 二面结果
            third_interview_result: 三面结果
            offer_status: OFFER状态
            page: 页码
            page_size: 每页数量

        Returns:
            包含候选人列表和分页信息的字典
        """
        from db.models import SalaryNegotiation, CandidateStageHistory
        from datetime import datetime, timedelta

        user = self.db.query(User).filter(User.id == user_id).first()
        if not user:
            raise ValueError("用户不存在")

        # 构建基础查询
        query = self.db.query(Candidate)

        # 根据用户角色过滤
        if user.role == UserRole.INTERVIEWER.value:
            # 从历史记录中查找关联的候选人
            related_candidate_ids = self.db.query(CandidateStageHistory.candidate_id).filter(
                or_(
                    CandidateStageHistory.stage_owner == user_id,
                    CandidateStageHistory.next_stage_owner == user_id
                )
            ).distinct().all()
            candidate_ids = [cid[0] for cid in related_candidate_ids]

            # 同时包含当前阶段负责人是该用户的候选人（如新添加尚未产生历史记录的）
            current_owner_ids = self.db.query(Candidate.id).filter(
                Candidate.current_stage_owner == user_id
            ).all()
            candidate_ids.extend([cid[0] for cid in current_owner_ids])
            candidate_ids = list(set(candidate_ids))

            if candidate_ids:
                query = query.filter(Candidate.id.in_(candidate_ids))
            else:
                query = query.filter(Candidate.id == -1)

        # JD筛选（支持多选）
        jd_id_list = [int(j.strip()) for j in jd_ids.split(',') if j.strip()] if jd_ids else []
        if jd_id and not jd_id_list:
            jd_id_list = [jd_id]
        if jd_id_list:
            query = query.filter(Candidate.jd_id.in_(jd_id_list))

        # 流程进展筛选（支持多选）
        stage_list = [s.strip() for s in stages.split(',') if s.strip()] if stages else []
        if stage and not stage_list:
            stage_list = [stage]
        if stage_list:
            query = query.filter(Candidate.current_stage.in_(stage_list))

        # 关键词搜索（候选人姓名或职位名称）
        if keyword:
            keyword_like = f"%{keyword}%"
            jd_ids_by_title = self.db.query(JobDescription.id).filter(
                JobDescription.job_title.like(keyword_like)
            ).all()
            jd_id_list = [j[0] for j in jd_ids_by_title]
            query = query.filter(
                or_(
                    Candidate.name.like(keyword_like),
                    Candidate.jd_id.in_(jd_id_list) if jd_id_list else False
                )
            )

        # 日期范围筛选
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                query = query.filter(Candidate.created_at >= start_dt)
            except ValueError:
                pass
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
                query = query.filter(Candidate.created_at < end_dt)
            except ValueError:
                pass

        # 部门筛选（支持多选，通过JD关联）
        dept_names = [d.strip() for d in departments.split(',') if d.strip()] if departments else []
        if department and not dept_names:
            dept_names = [department]
        if dept_names:
            dept_objs = self.db.query(DepartmentModel).filter(DepartmentModel.name.in_(dept_names)).all()
            if dept_objs:
                dept_ids = [d.id for d in dept_objs]
                jd_ids_by_dept = self.db.query(JobDescription.id).filter(
                    JobDescription.department_id.in_(dept_ids)
                ).all()
            else:
                jd_ids_by_dept = []
            dept_jd_ids = [j[0] for j in jd_ids_by_dept]
            if dept_jd_ids:
                query = query.filter(Candidate.jd_id.in_(dept_jd_ids))
            else:
                query = query.filter(Candidate.id == -1)

        # 简历筛选结果筛选（支持多选）
        sr_list = [s.strip() for s in screening_results.split(',') if s.strip()] if screening_results else []
        if screening_result and not sr_list:
            sr_list = [screening_result]
        if sr_list:
            sr_cid_sets = []
            for sr in sr_list:
                if sr == CandidateStageResult.PENDING.value:
                    pending_cids = self.db.query(Candidate.id).filter(
                        Candidate.current_stage == CandidateStage.RESUME_SCREENING.value,
                        Candidate.current_stage_result == CandidateStageResult.PENDING.value
                    ).all()
                    sr_cid_sets.extend([c[0] for c in pending_cids])
                else:
                    # 子查询：每个候选人在该阶段的最新历史记录ID
                    latest_history = select(
                        func.max(CandidateStageHistory.id)
                    ).filter(
                        CandidateStageHistory.stage == CandidateStage.RESUME_SCREENING.value,
                        or_(
                            CandidateStageHistory.is_abnormal_terminated == False,
                            CandidateStageHistory.is_abnormal_terminated == None
                        ),
                        or_(
                            CandidateStageHistory.stage_result != "转交",
                            CandidateStageHistory.stage_result == None
                        )
                    ).group_by(CandidateStageHistory.candidate_id)

                    cids = self.db.query(CandidateStageHistory.candidate_id).filter(
                        CandidateStageHistory.id.in_(latest_history),
                        CandidateStageHistory.stage_result == sr
                    ).all()
                    sr_cid_sets.extend([c[0] for c in cids])
            sr_cid_list = list(set(sr_cid_sets))
            query = query.filter(Candidate.id.in_(sr_cid_list)) if sr_cid_list else query.filter(Candidate.id == -1)

        # 一面结果筛选（支持多选）
        fi_list = [s.strip() for s in first_interview_results.split(',') if s.strip()] if first_interview_results else []
        if first_interview_result and not fi_list:
            fi_list = [first_interview_result]
        if fi_list:
            fi_cid_sets = []
            for fi in fi_list:
                if fi == CandidateStageResult.PENDING.value:
                    pending_cids = self.db.query(Candidate.id).filter(
                        Candidate.current_stage == CandidateStage.FIRST_INTERVIEW.value,
                        Candidate.current_stage_result == CandidateStageResult.PENDING.value
                    ).all()
                    fi_cid_sets.extend([c[0] for c in pending_cids])
                else:
                    # 子查询：每个候选人在一面阶段的最新历史记录ID
                    latest_history = select(
                        func.max(CandidateStageHistory.id)
                    ).filter(
                        CandidateStageHistory.stage == CandidateStage.FIRST_INTERVIEW.value,
                        or_(
                            CandidateStageHistory.is_abnormal_terminated == False,
                            CandidateStageHistory.is_abnormal_terminated == None
                        ),
                        or_(
                            CandidateStageHistory.stage_result != "转交",
                            CandidateStageHistory.stage_result == None
                        )
                    ).group_by(CandidateStageHistory.candidate_id)

                    cids = self.db.query(CandidateStageHistory.candidate_id).filter(
                        CandidateStageHistory.id.in_(latest_history),
                        CandidateStageHistory.stage_result == fi
                    ).all()
                    fi_cid_sets.extend([c[0] for c in cids])
            fi_cid_list = list(set(fi_cid_sets))
            query = query.filter(Candidate.id.in_(fi_cid_list)) if fi_cid_list else query.filter(Candidate.id == -1)

        # 二面结果筛选（支持多选）
        si_list = [s.strip() for s in second_interview_results.split(',') if s.strip()] if second_interview_results else []
        if second_interview_result and not si_list:
            si_list = [second_interview_result]
        if si_list:
            si_cid_sets = []
            for si in si_list:
                if si == CandidateStageResult.PENDING.value:
                    pending_cids = self.db.query(Candidate.id).filter(
                        Candidate.current_stage == CandidateStage.SECOND_INTERVIEW.value,
                        Candidate.current_stage_result == CandidateStageResult.PENDING.value
                    ).all()
                    si_cid_sets.extend([c[0] for c in pending_cids])
                else:
                    # 子查询：每个候选人在二面阶段的最新历史记录ID
                    latest_history = select(
                        func.max(CandidateStageHistory.id)
                    ).filter(
                        CandidateStageHistory.stage == CandidateStage.SECOND_INTERVIEW.value,
                        or_(
                            CandidateStageHistory.is_abnormal_terminated == False,
                            CandidateStageHistory.is_abnormal_terminated == None
                        ),
                        or_(
                            CandidateStageHistory.stage_result != "转交",
                            CandidateStageHistory.stage_result == None
                        )
                    ).group_by(CandidateStageHistory.candidate_id)

                    cids = self.db.query(CandidateStageHistory.candidate_id).filter(
                        CandidateStageHistory.id.in_(latest_history),
                        CandidateStageHistory.stage_result == si
                    ).all()
                    si_cid_sets.extend([c[0] for c in cids])
            si_cid_list = list(set(si_cid_sets))
            query = query.filter(Candidate.id.in_(si_cid_list)) if si_cid_list else query.filter(Candidate.id == -1)

        # 三面结果筛选（支持多选）
        ti_list = [s.strip() for s in third_interview_results.split(',') if s.strip()] if third_interview_results else []
        if third_interview_result and not ti_list:
            ti_list = [third_interview_result]
        if ti_list:
            ti_cid_sets = []
            for ti in ti_list:
                if ti == CandidateStageResult.PENDING.value:
                    pending_cids = self.db.query(Candidate.id).filter(
                        Candidate.current_stage == CandidateStage.THIRD_INTERVIEW.value,
                        Candidate.current_stage_result == CandidateStageResult.PENDING.value
                    ).all()
                    ti_cid_sets.extend([c[0] for c in pending_cids])
                else:
                    latest_history = select(
                        func.max(CandidateStageHistory.id)
                    ).filter(
                        CandidateStageHistory.stage == CandidateStage.THIRD_INTERVIEW.value,
                        or_(
                            CandidateStageHistory.is_abnormal_terminated == False,
                            CandidateStageHistory.is_abnormal_terminated == None
                        ),
                        or_(
                            CandidateStageHistory.stage_result != "转交",
                            CandidateStageHistory.stage_result == None
                        )
                    ).group_by(CandidateStageHistory.candidate_id)

                    cids = self.db.query(CandidateStageHistory.candidate_id).filter(
                        CandidateStageHistory.id.in_(latest_history),
                        CandidateStageHistory.stage_result == ti
                    ).all()
                    ti_cid_sets.extend([c[0] for c in cids])
            ti_cid_list = list(set(ti_cid_sets))
            query = query.filter(Candidate.id.in_(ti_cid_list)) if ti_cid_list else query.filter(Candidate.id == -1)

        # OFFER状态筛选（支持多选）
        os_list = [s.strip() for s in offer_statuses.split(',') if s.strip()] if offer_statuses else []
        if offer_status and not os_list:
            os_list = [offer_status]
        if os_list:
            latest_salary_negotiation = get_latest_salary_negotiation_subquery(self.db)
            cids = self.db.query(latest_salary_negotiation.c.candidate_id).filter(
                latest_salary_negotiation.c.offer_status.in_(os_list)
            ).all()
            cid_list = [c[0] for c in cids]
            query = query.filter(Candidate.id.in_(cid_list)) if cid_list else query.filter(Candidate.id == -1)

        # 计算总数
        total = query.count()

        # 分页（终止流程的候选人下沉到底部）
        from sqlalchemy import case
        terminated_order = case(
            (Candidate.current_stage == CandidateStage.TERMINATED.value, 1),
            else_=0
        )
        offset = (page - 1) * page_size
        candidates = query.order_by(terminated_order, Candidate.created_at.desc()).offset(offset).limit(page_size).all()

        # 格式化返回数据 - 包含所有需要的字段
        items = []
        for c in candidates:
            # 获取JD信息（部门和职位）
            jd = self.db.query(JobDescription).filter(JobDescription.id == c.jd_id).first()
            department = (jd.department_ref.name if jd.department_ref else jd.department) if jd else None
            job_title = jd.job_title if jd else None

            # 获取各阶段结果（按ID倒序，取每个阶段最新的一条）
            stage_histories = self.db.query(CandidateStageHistory).filter(
                CandidateStageHistory.candidate_id == c.id,
                or_(
                    CandidateStageHistory.is_abnormal_terminated == False,
                    CandidateStageHistory.is_abnormal_terminated == None
                ),
                or_(
                    CandidateStageHistory.stage_result != "转交",
                    CandidateStageHistory.stage_result == None
                )
            ).order_by(CandidateStageHistory.id.desc()).all()

            screening_result = None
            first_interview_result = None
            second_interview_result = None
            third_interview_result = None

            for history in stage_histories:
                if history.stage == CandidateStage.RESUME_SCREENING.value and screening_result is None:
                    screening_result = history.stage_result
                elif history.stage == CandidateStage.FIRST_INTERVIEW.value and first_interview_result is None:
                    first_interview_result = history.stage_result
                elif history.stage == CandidateStage.SECOND_INTERVIEW.value and second_interview_result is None:
                    second_interview_result = history.stage_result
                elif history.stage == CandidateStage.THIRD_INTERVIEW.value and third_interview_result is None:
                    third_interview_result = history.stage_result

            # 如果当前阶段是简历筛选且没有历史记录，使用current_stage_result
            if c.current_stage == CandidateStage.RESUME_SCREENING.value and screening_result is None:
                screening_result = c.current_stage_result
            # 如果当前阶段是一面且没有历史记录，使用current_stage_result
            if c.current_stage == CandidateStage.FIRST_INTERVIEW.value and first_interview_result is None:
                first_interview_result = c.current_stage_result
            # 如果当前阶段是二面且没有历史记录，使用current_stage_result
            if c.current_stage == CandidateStage.SECOND_INTERVIEW.value and second_interview_result is None:
                second_interview_result = c.current_stage_result
            # 如果当前阶段是三面且没有历史记录，使用current_stage_result
            if c.current_stage == CandidateStage.THIRD_INTERVIEW.value and third_interview_result is None:
                third_interview_result = c.current_stage_result

            # 获取OFFER状态和入职状态
            salary_neg = get_latest_salary_negotiation_record(self.db, c.id)
            offer_status = salary_neg.offer_status if salary_neg else None
            is_onboarded = salary_neg.is_onboarded if salary_neg else False

            items.append({
                "id": c.id,
                "candidate_number": c.candidate_number,
                "name": c.name,
                "jd_id": c.jd_id,
                "department": department,
                "job_title": job_title,
                "current_stage": c.current_stage,
                "current_stage_result": c.current_stage_result,
                "current_stage_owner_name": (
                    c.stage_owner.real_name
                    if c.current_stage != CandidateStage.TERMINATED.value and c.stage_owner
                    else None
                ),
                "screening_result": screening_result,
                "first_interview_result": first_interview_result,
                "second_interview_result": second_interview_result,
                "third_interview_result": third_interview_result,
                "offer_status": offer_status,
                "is_onboarded": is_onboarded,
                "is_parsed": c.is_parsed,
                "ai_score_total": c.ai_score_total,
                "hard_requirements_passed": c.hard_requirements_passed,
                "resume_file_path": c.resume_file_path,
                "created_at": c.created_at.isoformat() if c.created_at else None
            })

        result = {
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size,
            "items": items
        }

        return result

    def export_candidates_to_excel(
        self,
        user_id: int,
        jd_id: Optional[int] = None,
        stage: Optional[str] = None,
        candidate_ids: Optional[List[int]] = None
    ) -> bytes:
        """
        导出候选人列表为Excel

        Args:
            user_id: 用户ID
            jd_id: JD ID（可选）
            stage: 阶段（可选）
            candidate_ids: 指定导出的候选人ID列表（可选，优先级最高）

        Returns:
            Excel文件的字节内容
        """
        import io
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from sqlalchemy import func, and_
        from db.models import SalaryNegotiation, ResumeEvaluationRule

        # 获取候选人数据（不分页，获取全部）
        query = self.db.query(Candidate)

        # 如果指定了候选人ID列表，优先使用
        if candidate_ids:
            query = query.filter(Candidate.id.in_(candidate_ids))
        else:
            if jd_id:
                query = query.filter(Candidate.jd_id == jd_id)
            if stage:
                query = query.filter(Candidate.current_stage == stage)

        candidates = query.order_by(Candidate.created_at.desc()).all()

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "候选人列表"

        # 定义表头
        headers = [
            "序号", "候选人姓名", "所属部门", "应聘职位", "当前负责人", "当前流程",
            "简历筛选", "一面", "二面", "三面", "OFFER状态", "入职",
            "AI简历评分", "简历添加时间"
        ]

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row_idx, candidate in enumerate(candidates, 2):
            # 获取JD信息
            jd = self.db.query(JobDescription).filter(JobDescription.id == candidate.jd_id).first()
            department = (jd.department_ref.name if jd.department_ref else jd.department) if jd else "-"
            job_title = jd.job_title if jd else "-"

            # 获取各阶段结果（按ID倒序，取每个阶段最新的一条）
            stage_histories = self.db.query(CandidateStageHistory).filter(
                CandidateStageHistory.candidate_id == candidate.id,
                or_(
                    CandidateStageHistory.is_abnormal_terminated == False,
                    CandidateStageHistory.is_abnormal_terminated == None
                ),
                or_(
                    CandidateStageHistory.stage_result != "转交",
                    CandidateStageHistory.stage_result == None
                )
            ).order_by(CandidateStageHistory.id.desc()).all()

            screening_result = "-"
            first_interview_result = "-"
            second_interview_result = "-"
            third_interview_result = "-"

            for history in stage_histories:
                if history.stage == CandidateStage.RESUME_SCREENING.value and screening_result == "-":
                    screening_result = history.stage_result or "-"
                elif history.stage == CandidateStage.FIRST_INTERVIEW.value and first_interview_result == "-":
                    first_interview_result = history.stage_result or "-"
                elif history.stage == CandidateStage.SECOND_INTERVIEW.value and second_interview_result == "-":
                    second_interview_result = history.stage_result or "-"
                elif history.stage == CandidateStage.THIRD_INTERVIEW.value and third_interview_result == "-":
                    third_interview_result = history.stage_result or "-"

            # 获取OFFER状态和入职状态（从SalaryNegotiation表）
            salary_neg = get_latest_salary_negotiation_record(self.db, candidate.id)
            offer_status = salary_neg.offer_status if salary_neg else "-"
            is_onboarded = salary_neg.is_onboarded if salary_neg else False

            # 计算AI简历评分满分
            total_score_max = 100  # 默认值
            if jd and jd.resume_rule_set_id:
                # 计算主要分满分（非加分项）- 按指标名称去重
                main_indicators = self.db.query(
                    ResumeEvaluationRule.indicator_name,
                    func.max(ResumeEvaluationRule.total_score).label('max_score')
                ).filter(
                    and_(
                        ResumeEvaluationRule.rule_set_id == jd.resume_rule_set_id,
                        ResumeEvaluationRule.is_bonus == False
                    )
                ).group_by(ResumeEvaluationRule.indicator_name).all()

                main_score_max = sum(ind.max_score or 0 for ind in main_indicators) if main_indicators else 0

                # 计算加分项满分 - 按指标名称去重
                bonus_indicators = self.db.query(
                    ResumeEvaluationRule.indicator_name,
                    func.max(ResumeEvaluationRule.total_score).label('max_score')
                ).filter(
                    and_(
                        ResumeEvaluationRule.rule_set_id == jd.resume_rule_set_id,
                        ResumeEvaluationRule.is_bonus == True
                    )
                ).group_by(ResumeEvaluationRule.indicator_name).all()

                bonus_score_max = sum(ind.max_score or 0 for ind in bonus_indicators) if bonus_indicators else 0

                total_score_max = main_score_max + bonus_score_max if (main_score_max + bonus_score_max) > 0 else 100

            # AI评分显示
            ai_score_display = f"{candidate.ai_score_total}/{total_score_max}" if candidate.ai_score_total else "-"

            # 获取当前负责人姓名
            owner_name = "-"
            if candidate.current_stage != CandidateStage.TERMINATED.value and candidate.current_stage_owner:
                owner = self.db.query(User).filter(User.id == candidate.current_stage_owner).first()
                if owner:
                    owner_name = owner.real_name or owner.username

            # 写入行数据
            row_data = [
                row_idx - 1,  # 序号
                candidate.name or "-",
                department,
                job_title,
                owner_name,
                candidate.current_stage or "-",
                screening_result,
                first_interview_result,
                second_interview_result,
                third_interview_result,
                offer_status or "-",
                "是" if is_onboarded else "否",
                ai_score_display,
                candidate.created_at.strftime("%Y-%m-%d %H:%M") if candidate.created_at else "-"
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        # 调整列宽
        column_widths = [8, 15, 15, 18, 12, 12, 10, 10, 10, 10, 12, 8, 15, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()
